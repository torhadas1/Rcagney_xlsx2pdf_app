import streamlit as st
import pandas as pd
import openpyxl
import win32com.client
import pythoncom
import io
import base64

def search_input(column, df):
    search_bar = st.text_input(str(column)) # creating a streamlit text input widget
    filtered_df = df[df[column].str.contains(search_bar)] # filtering a df by the text input inside the widget
    return filtered_df # returning the filtered df


def update_cell_value(input_file, sheet_name, cell_address, new_value):
    # Ensure the input file is an xlsx file
    if not input_file.lower().endswith(".xlsx"):
        raise ValueError("Invalid input file. Please provide a .xlsx file.")

    # Load the workbook
    wb = openpyxl.load_workbook(input_file)

    # Check if the sheet exists
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")

    # Access the sheet and cell
    sheet = wb[sheet_name]
    cell = sheet[cell_address]

    # Update the cell value
    cell.value = new_value

  # Reset the cursor to the beginning of the BytesIO object
    xlsx_data = io.BytesIO()
    wb.save(input_file)

    return xlsx_data


def xlsx2pdf(xlsx_path, output_path, sheet1, sheet2):
        # Ensure the input file is an xlsx file
    if not xlsx_path.lower().endswith(".xlsx"):
        raise ValueError("Invalid input file. Please provide a .xlsx file.")

    pythoncom.CoInitialize()      
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    # Read Excel File
    sheets = excel.Workbooks.Open(xlsx_path)
    work_sheets1 = sheets.Worksheets[sheet1]
    work_sheets2 = sheets.Worksheets[sheet2]
    # Convert into PDF File
    output_file = xlsx_path.replace('.xlsx','.pdf') 
    work_sheets1.ExportAsFixedFormat(0, output_path + str(sheet1) + '.pdf')
    work_sheets2.ExportAsFixedFormat(0, output_path + str(sheet2) + '.pdf')
    sheets.Close(False)
    excel.Quit()
    return output_path + str(sheet1) + '.pdf', output_path + str(sheet2) + '.pdf'
    

def display_pdf(pdf_file):
    with open(pdf_file, "rb") as f:
        pdf_data = f.read()

    # Encode the PDF data to base64
    pdf_b64 = base64.b64encode(pdf_data).decode("utf-8")

    # Display the PDF on Streamlit
    st.markdown(
        f'<embed src="data:application/pdf;base64,{pdf_b64}" width="600" height="400" type="application/pdf">',
        unsafe_allow_html=True,) # show the pdf as an iframe inside the app


st.set_page_config(layout= 'wide') # configuring the streamlit app page layout

col1, col2, col3, col4, col5 = st.columns([2,1,1,1,1]) # creating five columns the top of the page

with col1: # write a logo on col 1
    st.image(r'img/logo.png',width=350)


df = pd.read_excel('HCAD Data Source of Truth_commercial.xlsx' , sheet_name='Cost' , skiprows=3, converters={'ACCOUNT':str}) # reading the database into a df
df = df.fillna('')
df['ACCOUNT'] = df['ACCOUNT'].astype(str)

with col2:
    st.markdown('')
    st.markdown('')
    generator = st.button('Generate PDF', type='primary')

with col3:
    account = st.text_input('',key='account', placeholder ='Account#')

with col4:
    name = st.text_input('',key='name', placeholder ='Owner Name')

with col5:
    address = st.text_input('',key='address', placeholder ='Property Address')



if account != '':
    df = df[df['ACCOUNT'].str.contains(account, na=False)]
elif name != '':
    df = df[df['OWNER_NAME'].str.contains(name, na=False)]
elif address != '':
    df = df[df['PROPERTY ADDRESS'].str.contains(address, na=False)]


st.dataframe(df[['ACCOUNT','OWNER_NAME','PROPERTY ADDRESS','MARKET_AREA_1_DSCR']], width = 1000) # presenting the filtered df

st.markdown('---') # draw a divider between the sections

path_df = pd.read_csv('paths.csv') # read a csv file of the paths of the calculator xlsx and the output for the pdf
xlsx_path = path_df['paths'][0] # xlsx calculator file path
pdf_path = path_df['paths'][1] # the folder for the pdf paths


if generator:
    with st.spinner('Generating PDF'): # a spinner telling genarating pdf, for the process part
        if account is not None:
            update_cell_value('Commercial Cost Test Template.xlsx','SELECTION','B4',account) # updating the calculator calue with the account input

            comparision_pdf, HCAD_pdf = xlsx2pdf(xlsx_path,pdf_path,'Comparision','HCAD Cost') # creating two pdf's one for every sheet
            def xlsx2byte(path,sheet2remove):
                wb = openpyxl.load_workbook(path)
                xlsx_data = io.BytesIO()
                del wb[sheet2remove[0]]
                del wb[sheet2remove[1]]
                del wb[sheet2remove[2]]
                wb.save(xlsx_data)
                return xlsx_data
            HCAD_data = xlsx2byte('Commercial Cost Test Template.xlsx',['Nettles Cost','Comparision','List'])
            comparision_data = xlsx2byte('Commercial Cost Test Template.xlsx',['Nettles Cost','HCAD Cost','List'])

            bottom_col_left, bottom_col_right = st.columns(2) # creating two columns if the bottom of the page

            with bottom_col_left:
                bottom_col_left_left, bottom_col_left_right = st.columns([3,1])
                with bottom_col_left_left:
                    st.write('**Harris Country Account Information**') # a title with the name of the corresponding document
                with bottom_col_left_right:
                    st.download_button('Download xlsx',HCAD_data,'HCAD Calculator.xlsx')
                display_pdf(HCAD_pdf) # showing the pdf

            with bottom_col_right:
                bottom_col_right_left, bottom_col_right_right = st.columns([3,1])
                with bottom_col_right_left:
                    st.write('**ValueCloud Account Information**') # a title with the name of the corresponding document
                with bottom_col_right_right:
                    st.download_button('Download xlsx',comparision_data,'Comparision Calculator.xlsx')
                display_pdf(comparision_pdf)    # showing the pdf


